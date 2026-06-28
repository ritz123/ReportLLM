"""Parse uploaded expense sheets (.xlsx, .xls, .md)."""

import io
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from analyze_x import parse_md, parse_month_sheet, compute_observations, finalize_months  # noqa: E402
from data_loader import months_to_dataset  # noqa: E402


def _cell_str(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M:%S") if hasattr(value, "hour") else value.strftime("%Y-%m-%d")
    return str(value).strip()


def parse_xlsx_bytes(data: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl is required for Excel upload. Run: uv pip install openpyxl") from e

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    months = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                if not any(c is not None and str(c).strip() for c in row):
                    continue
                rows.append([_cell_str(c) for c in row])
            month = parse_month_sheet(sheet_name, rows)
            if month:
                months.append(month)
    finally:
        wb.close()

    if not months:
        raise ValueError("No monthly sheets found. Each sheet should be named like 'Jan 2022' or 'OCT 2021'.")

    return finalize_months(months)


def parse_upload(filename: str, data: bytes) -> dict:
    """Parse uploaded file and return full dataset dict."""
    name = filename.lower()
    if name.endswith(".md"):
        with tempfile.NamedTemporaryFile(suffix=".md", delete=False) as tmp:
            tmp.write(data)
            tmp_path = Path(tmp.name)
        try:
            months = parse_md(tmp_path)
        finally:
            tmp_path.unlink(missing_ok=True)
    elif name.endswith((".xlsx", ".xlsm", ".xltx")):
        months = parse_xlsx_bytes(data)
    elif name.endswith(".xls"):
        try:
            import xlrd
        except ImportError as e:
            raise RuntimeError("xlrd is required for .xls files. Prefer .xlsx or install xlrd.") from e
        with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
            tmp.write(data)
            tmp_path = Path(tmp.name)
        try:
            book = xlrd.open_workbook(tmp_path)
            months = []
            for sheet_name in book.sheet_names():
                sh = book.sheet_by_name(sheet_name)
                rows = []
                for rx in range(sh.nrows):
                    row = [_cell_str(sh.cell_value(rx, cx)) for cx in range(sh.ncols)]
                    if any(row):
                        rows.append(row)
                month = parse_month_sheet(sheet_name, rows)
                if month:
                    months.append(month)
            months = finalize_months(months)
        finally:
            tmp_path.unlink(missing_ok=True)
        if not months:
            raise ValueError("No monthly sheets found in .xls file.")
    else:
        raise ValueError(f"Unsupported file type: {filename}. Upload .xlsx, .xls, or .md")

    observations = compute_observations(months)
    dataset = months_to_dataset(months, observations)
    dataset["source_file"] = filename
    return dataset
